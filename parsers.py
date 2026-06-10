"""
parsers.py
----------
Turns heterogeneous log files into a uniform stream of `Record` objects so the
detection engine doesn't care about file format.

A Record has:
    file        - source file name
    location    - human readable position (line N / row N / Sheet!row)
    text        - the text to scan
    timestamp   - best-effort timestamp extracted from the record ("" if none)

Supported: .log .txt .out .json .jsonl .ndjson .csv .tsv .xlsx .xls
Anything else is read as plain text.
"""

from __future__ import annotations

import csv
import json
import os
import re
from dataclasses import dataclass
from typing import Iterator


@dataclass
class Record:
    file: str
    location: str
    text: str
    timestamp: str = ""


# Common timestamp shapes found across app/API/web-server logs.
_TS_PATTERNS = [
    r"\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:Z|[+\-]\d{2}:?\d{2})?",  # ISO 8601
    r"\d{2}/[A-Za-z]{3}/\d{4}:\d{2}:\d{2}:\d{2}\s?[+\-]\d{4}",                     # Apache/CLF
    r"[A-Za-z]{3}\s+\d{1,2}\s+\d{2}:\d{2}:\d{2}",                                  # syslog
    r"\d{2}[-/]\d{2}[-/]\d{4}[ T]\d{2}:\d{2}:\d{2}",                               # dd-mm-yyyy hh:mm:ss
]
_TS_RE = re.compile("|".join(f"(?:{p})" for p in _TS_PATTERNS))

# Field names that typically hold a timestamp in structured logs.
_TS_FIELDS = ("timestamp", "time", "ts", "@timestamp", "date", "datetime",
              "eventtime", "logtime", "created_at", "createdat")


def extract_timestamp(text: str) -> str:
    m = _TS_RE.search(text)
    return m.group(0) if m else ""


def _ts_from_obj(obj) -> str:
    if isinstance(obj, dict):
        lowered = {str(k).lower(): v for k, v in obj.items()}
        for key in _TS_FIELDS:
            if key in lowered and lowered[key] not in (None, ""):
                return str(lowered[key])
    return ""


# --------------------------------------------------------------------------- #
# Per-format parsers
# --------------------------------------------------------------------------- #

def _parse_text(path: str, name: str) -> Iterator[Record]:
    with open(path, "r", encoding="utf-8", errors="replace") as fh:
        for i, line in enumerate(fh, 1):
            line = line.rstrip("\n")
            if not line.strip():
                continue
            yield Record(name, f"line {i}", line, extract_timestamp(line))


def _parse_jsonl(path: str, name: str) -> Iterator[Record]:
    with open(path, "r", encoding="utf-8", errors="replace") as fh:
        for i, line in enumerate(fh, 1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                ts = _ts_from_obj(obj) or extract_timestamp(line)
                text = json.dumps(obj, ensure_ascii=False)
            except json.JSONDecodeError:
                ts, text = extract_timestamp(line), line
            yield Record(name, f"line {i}", text, ts)


def _walk_json_records(obj):
    """Yield (index, record_obj) for the most natural 'record' granularity."""
    if isinstance(obj, list):
        for i, item in enumerate(obj, 1):
            yield i, item
    elif isinstance(obj, dict):
        # A dict whose values are all lists/dicts of records -> flatten one level
        list_vals = [v for v in obj.values() if isinstance(v, list)]
        if len(list_vals) == 1 and len(obj) == 1:
            for i, item in enumerate(list_vals[0], 1):
                yield i, item
        else:
            yield 1, obj
    else:
        yield 1, obj


def _parse_json(path: str, name: str) -> Iterator[Record]:
    with open(path, "r", encoding="utf-8", errors="replace") as fh:
        content = fh.read()
    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        # Not a single JSON document - fall back to line scanning (maybe JSONL)
        yield from _parse_jsonl(path, name)
        return
    for idx, rec in _walk_json_records(data):
        text = json.dumps(rec, ensure_ascii=False)
        ts = _ts_from_obj(rec) or extract_timestamp(text)
        yield Record(name, f"record {idx}", text, ts)


def _parse_csv(path: str, name: str, delimiter: str = ",") -> Iterator[Record]:
    with open(path, "r", encoding="utf-8", errors="replace", newline="") as fh:
        reader = csv.reader(fh, delimiter=delimiter)
        header = None
        ts_idx = None
        for i, row in enumerate(reader, 1):
            if i == 1:
                header = [c.strip().lower() for c in row]
                for j, col in enumerate(header):
                    if col in _TS_FIELDS:
                        ts_idx = j
                        break
                continue
            if not any(cell.strip() for cell in row):
                continue
            text = delimiter.join(row)
            ts = ""
            if ts_idx is not None and ts_idx < len(row):
                ts = row[ts_idx]
            if not ts:
                ts = extract_timestamp(text)
            yield Record(name, f"row {i}", text, ts)


def _parse_xlsx(path: str, name: str) -> Iterator[Record]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.worksheets:
        header = None
        ts_idx = None
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            cells = ["" if c is None else str(c) for c in row]
            if i == 1:
                header = [c.strip().lower() for c in cells]
                for j, col in enumerate(header):
                    if col in _TS_FIELDS:
                        ts_idx = j
                        break
                # Heuristic: if first row doesn't look like a header, still scan it
                if not any(h.isalpha() for h in "".join(header)):
                    ts_idx = None
                    header = None
                else:
                    continue
            if not any(c.strip() for c in cells):
                continue
            text = " | ".join(cells)
            ts = ""
            if ts_idx is not None and ts_idx < len(cells):
                ts = cells[ts_idx]
            if not ts:
                ts = extract_timestamp(text)
            yield Record(name, f"{sheet.title}!row {i}", text, ts)
    wb.close()


# --------------------------------------------------------------------------- #
# Dispatcher
# --------------------------------------------------------------------------- #

_DISPATCH = {
    ".json": _parse_json,
    ".jsonl": _parse_jsonl,
    ".ndjson": _parse_jsonl,
    ".csv": lambda p, n: _parse_csv(p, n, ","),
    ".tsv": lambda p, n: _parse_csv(p, n, "\t"),
    ".xlsx": _parse_xlsx,
    ".xlsm": _parse_xlsx,
    ".xls": _parse_xlsx,
}


def parse_file(path: str) -> Iterator[Record]:
    """Yield Records for a single file, dispatching on extension."""
    name = os.path.basename(path)
    ext = os.path.splitext(path)[1].lower()
    parser = _DISPATCH.get(ext, _parse_text)
    try:
        yield from parser(path, name)
    except Exception as exc:  # never let one bad file kill the whole run
        yield Record(name, "n/a", f"[PARSER ERROR: {exc}]", "")


def gather_files(inputs: list[str]) -> list[str]:
    """Expand directories into a flat list of files."""
    files: list[str] = []
    for item in inputs:
        if os.path.isdir(item):
            for root, _dirs, names in os.walk(item):
                for n in names:
                    files.append(os.path.join(root, n))
        elif os.path.isfile(item):
            files.append(item)
    return sorted(set(files))
